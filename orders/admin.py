from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.http import HttpResponse
from django import forms
from django.utils import timezone
from django.utils.html import format_html
from .models import (
    Employee, Restaurant, MenuItemGroup, MenuItem, MenuItemImage,
    Order, OrderItem, BalanceTransaction, WorkDayCalendar, Settings, ProductCategory, SystemConfig, GlobalWorkDay
)
import csv
import io
import zipfile
import tempfile
import os
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta
from django.core.files import File


class ExportCsvMixin:
    """Миксин для экспорта выбранных записей в CSV"""
    
    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural}.csv'
        response.write('\ufeff')  # BOM for Excel UTF-8
        
        writer = csv.writer(response)
        
        # Write header with verbose names
        header = []
        for field in meta.fields:
            header.append(field.verbose_name if hasattr(field, 'verbose_name') else field.name)
        writer.writerow(header)
        
        # Write data rows
        for obj in queryset:
            row = []
            for field in field_names:
                value = getattr(obj, field)
                if callable(value):
                    value = value()
                if value is None:
                    value = ''
                row.append(str(value))
            writer.writerow(row)
        
        return response
    
    export_as_csv.short_description = "Экспорт выбранных в CSV"



class ExcelImportForm(forms.Form):
    """Форма для импорта Excel с группами товаров"""
    excel_file = forms.FileField(label='Excel файл (.xlsx)')
    restaurant_name = forms.CharField(
        label='Название ресторана (по умолчанию)',
        initial='ВкусВилл',
        required=False,
        help_text='Название ресторана, если не указано в файле'
    )
    date_column = forms.CharField(
        label='Столбец с датой',
        required=False,
        initial='дата',
        help_text='Название столбца с датой (если есть). Если не указан, дата будет взята из названия листа'
    )
    name_column = forms.CharField(
        label='Столбец "Название"',
        initial='Наименование',
        required=False,
        help_text='Название столбца с названием товара (оставьте пустым для автоматического определения)'
    )
    price_column = forms.CharField(
        label='Столбец "Цена"',
        required=False,
        initial='Цена',
        help_text='Название столбца с ценой (оставьте пустым для автоматического определения)'
    )
    category_column = forms.CharField(
        label='Столбец "Категория"',
        required=False,
        initial='Категория',
        help_text='Название столбца с категорией товара'
    )
    restaurant_column = forms.CharField(
        label='Столбец "Ресторан"',
        required=False,
        initial='Ресторан',
        help_text='Название столбца с рестораном'
    )
    description_column = forms.CharField(
        label='Столбец "Описание"',
        required=False,
        initial='описание',
        help_text='Название столбца с описанием (необязательно)'
    )


class CSVImportForm(forms.Form):
    """Форма для импорта CSV с маппингом столбцов"""
    csv_file = forms.FileField(label='CSV файл')
    zip_file = forms.FileField(
        label='ZIP архив с изображениями',
        required=False,
        help_text='ZIP архив с фотографиями товаров (необязательно). В CSV укажите путь к файлу относительно корня архива, например: photos_today/img1.png'
    )
    
    # Маппинг столбцов CSV на поля модели
    name_column = forms.CharField(
        label='Столбец "Название"',
        initial='название',
        help_text='Название столбца в CSV для поля "Название товара"'
    )
    group_column = forms.CharField(
        label='Столбец "Группа"',
        initial='группа',
        help_text='Название столбца в CSV для поля "Группа товара"'
    )
    restaurant_column = forms.CharField(
        label='Столбец "Ресторан"',
        initial='ресторан',
        help_text='Название столбца в CSV для поля "Ресторан"'
    )
    price_column = forms.CharField(
        label='Столбец "Цена"',
        required=False,
        initial='цена',
        help_text='Название столбца в CSV для поля "Цена" (необязательно)'
    )
    description_column = forms.CharField(
        label='Столбец "Описание"',
        required=False,
        initial='описание',
        help_text='Название столбца в CSV для поля "Описание" (необязательно)'
    )
    image_column = forms.CharField(
        label='Столбец "Изображение"',
        required=False,
        initial='изображение',
        help_text='Название столбца в CSV для пути к изображению в ZIP архиве (необязательно). Например: photos_today/img1.png'
    )
    
    encoding = forms.ChoiceField(
        label='Кодировка файла',
        choices=[
            ('utf-8', 'UTF-8'),
            ('windows-1251', 'Windows-1251'),
            ('cp1251', 'CP1251'),
        ],
        initial='utf-8'
    )
    
    delimiter = forms.ChoiceField(
        label='Разделитель',
        choices=[
            (';', 'Точка с запятой (;)'),
            (',', 'Запятая (,)'),
            ('\t', 'Табуляция'),
        ],
        initial=';'
    )


class BalanceTransactionInline(admin.TabularInline):
    """История операций с балансом в карточке сотрудника"""
    model = BalanceTransaction
    extra = 0
    readonly_fields = ['transaction_type', 'amount', 'balance_after', 'order', 'comment', 'created_at']
    can_delete = False
    fields = ['created_at', 'transaction_type', 'amount', 'balance_after', 'order', 'comment']
    ordering = ['-created_at']
    
    def has_add_permission(self, request, obj=None):
        return False


class OrderInline(admin.TabularInline):
    """История заказов в карточке сотрудника"""
    model = Order
    extra = 0
    readonly_fields = ['order_date', 'group', 'total_amount', 'status', 'created_at']
    can_delete = False
    fields = ['order_date', 'group', 'total_amount', 'status', 'created_at']
    ordering = ['-order_date', '-created_at']
    show_change_link = True
    
    def has_add_permission(self, request, obj=None):
        return False


@admin.register(Employee)
class EmployeeAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['name', 'email', 'get_username', 'telegram_id', 'balance', 'personal_balance', 'finance_link', 'is_active', 'is_approved', 'min_balance_limit', 'daily_balance_amount']
    list_filter = ['is_active', 'is_approved']
    search_fields = ['user__username', 'user__email', 'user__first_name', 'user__last_name', 'telegram_id']
    
    def get_username(self, obj):
        """Получить username из связанного User"""
        return obj.user.username if obj.user else '-'
    get_username.short_description = 'Логин'
    fieldsets = (
        ('Учетная запись Django', {
            'fields': ('user',),
            'description': 'Связь с пользователем Django для аутентификации. При создании сотрудника можно создать нового пользователя или выбрать существующего.'
        }),
        ('Основная информация', {
            'fields': ('telegram_id',)
        }),
        ('Баланс и настройки', {
            'fields': ('balance', 'personal_balance', 'daily_balance_amount', 'min_balance_limit')
        }),
        ('Статус', {
            'fields': ('is_active', 'is_approved')
        }),
        ('Системная информация', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at', 'name', 'email']
    
    def save_model(self, request, obj, form, change):
        """Создание/обновление связанного User при сохранении Employee"""
        from django.contrib.auth.models import User
        
        # Если Employee не связан с User, создаем нового пользователя
        if not obj.user_id:
            # Используем email как username, если он есть
            username = obj.email if hasattr(obj, 'email') and obj.email else f"employee_{obj.id or 'new'}"
            # Убеждаемся, что username уникален
            base_username = username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}_{counter}"
                counter += 1
            
            # Создаем пользователя
            user = User.objects.create_user(
                username=username,
                email=getattr(obj, 'email', '') or '',
            )
            obj.user = user
        
        super().save_model(request, obj, form, change)
    inlines = [BalanceTransactionInline, OrderInline]
    actions = ['approve_accounts', 'unapprove_accounts']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:employee_id>/finance/',
                self.admin_site.admin_view(self.finance_settings_view),
                name='orders_employee_finance'
            ),
        ]
        return custom_urls + urls

    def finance_link(self, obj):
        return format_html('<a href="{}">Финансы</a>', f'{obj.id}/finance/')
    finance_link.short_description = 'Настройка финансов'

    def finance_settings_view(self, request, employee_id):
        employee = Employee.objects.filter(id=employee_id).first()
        if not employee:
            self.message_user(request, 'Сотрудник не найден', level='error')
            return redirect('admin:orders_employee_changelist')

        today = timezone.now().date()
        day_record = WorkDayCalendar.objects.filter(employee=employee, date=today).first()
        is_sick_today = bool(day_record and day_record.day_type == 'sick_leave')

        if request.method == 'POST':
            action_name = request.POST.get('action_name')
            try:
                if action_name == 'save_settings':
                    employee.balance = Decimal(request.POST.get('balance', employee.balance))
                    employee.personal_balance = Decimal(request.POST.get('personal_balance', employee.personal_balance))
                    employee.daily_balance_amount = Decimal(request.POST.get('daily_balance_amount', employee.daily_balance_amount))
                    employee.min_balance_limit = Decimal(request.POST.get('min_balance_limit', employee.min_balance_limit))
                    employee.save()
                    self.message_user(request, 'Финансовые настройки сохранены', level='success')

                elif action_name == 'deposit_personal':
                    deposit = Decimal(request.POST.get('personal_deposit', '0') or '0')
                    if deposit <= 0:
                        self.message_user(request, 'Депозит должен быть больше 0', level='error')
                    else:
                        employee.personal_balance += deposit
                        employee.save()
                        self.message_user(
                            request,
                            f'Личный баланс пополнен на {deposit} ₽ (текущий: {employee.personal_balance} ₽)',
                            level='success'
                        )

                elif action_name == 'accrual_company':
                    amount = employee.daily_balance_amount
                    employee.balance += amount
                    employee.save()
                    BalanceTransaction.objects.create(
                        employee=employee,
                        transaction_type='accrual',
                        amount=amount,
                        balance_after=employee.balance,
                        comment=f'Ручное начисление компании через страницу финансов ({today})'
                    )
                    self.message_user(request, f'Начислено от фирмы: {amount} ₽', level='success')

                elif action_name == 'toggle_sick_today':
                    mark_sick = request.POST.get('is_sick_today') == 'on'
                    if mark_sick:
                        WorkDayCalendar.objects.update_or_create(
                            employee=employee,
                            date=today,
                            defaults={'day_type': 'sick_leave', 'comment': 'Отмечено в админке: болеет'}
                        )
                    else:
                        WorkDayCalendar.objects.update_or_create(
                            employee=employee,
                            date=today,
                            defaults={'day_type': 'work', 'comment': 'Отмечено в админке: рабочий день'}
                        )
                    self.message_user(request, 'Отметка по болезни на сегодня сохранена', level='success')

            except Exception as e:
                self.message_user(request, f'Ошибка: {e}', level='error')

            return redirect(request.path)

        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'title': f'Финансы сотрудника: {employee.name}',
            'employee': employee,
            'is_sick_today': is_sick_today,
            'today': today,
        }
        return render(request, 'admin/orders/employee/finance_settings.html', context)
    
    def approve_accounts(self, request, queryset):
        """Подтвердить учётные записи"""
        updated = queryset.update(is_approved=True)
        self.message_user(request, f'Подтверждено учётных записей: {updated}')
    approve_accounts.short_description = 'Подтвердить выбранные учётные записи'
    
    def unapprove_accounts(self, request, queryset):
        """Отменить подтверждение учётных записей"""
        updated = queryset.update(is_approved=False)
        self.message_user(request, f'Отменено подтверждение учётных записей: {updated}')
    unapprove_accounts.short_description = 'Отменить подтверждение выбранных учётных записей'


@admin.register(Restaurant)
class RestaurantAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['name', 'is_active', 'period_start', 'period_end']
    list_filter = ['is_active']
    search_fields = ['name']


@admin.register(MenuItemGroup)
class MenuItemGroupAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['name', 'order', 'is_active', 'period_start', 'period_end', 'is_selection_closed', 'created_at']
    list_filter = ['is_active', 'is_selection_closed']
    search_fields = ['name', 'description']
    ordering = ['order', 'name']
    change_list_template = 'admin/orders/menuitemgroup/change_list.html'
    fieldsets = (
        ('Основная информация', {
            'fields': ('name', 'description', 'order')
        }),
        ('Период видимости', {
            'fields': ('period_start', 'period_end')
        }),
        ('Статус', {
            'fields': ('is_active', 'is_selection_closed')
        }),
        ('Системная информация', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at']
    actions = ['close_selection', 'open_selection', 'export_orders_report']
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='orders_menuitemgroup_import_excel'),
            path('export-template/', self.admin_site.admin_view(self.export_template), name='orders_menuitemgroup_export_template'),
        ]
        return custom_urls + urls
    
    def close_selection(self, request, queryset):
        """Закрыть выбор для групп и финализировать заказы (списать средства)"""
        from django.db import transaction
        
        total_finalized = 0
        for group in queryset:
            if not group.is_selection_closed:
                finalized = group.close_selection_and_finalize_orders()
                total_finalized += finalized
        
        self.message_user(
            request, 
            f'Закрыт выбор для {queryset.count()} групп. Финализировано заказов: {total_finalized}',
            level='success'
        )
    close_selection.short_description = 'Закрыть выбор и финализировать заказы (списать средства)'
    
    def open_selection(self, request, queryset):
        """Открыть выбор для групп"""
        updated = queryset.update(is_selection_closed=False)
        self.message_user(request, f'Открыт выбор для групп: {updated}')
    open_selection.short_description = 'Открыть выбор для выбранных групп'
    
    def import_excel(self, request):
        """Импорт групп товаров из Excel файла (каждый лист = группа = меню на день)"""
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                default_restaurant_name = form.cleaned_data['restaurant_name']
                
                # Маппинг столбцов (используем значения из формы, если указаны)
                column_mapping = {
                    'date': form.cleaned_data.get('date_column', '').strip() if form.cleaned_data.get('date_column') else None,
                    'name': form.cleaned_data.get('name_column', '').strip() if form.cleaned_data.get('name_column') else None,
                    'price': form.cleaned_data.get('price_column', '').strip() if form.cleaned_data.get('price_column') else None,
                    'description': form.cleaned_data.get('description_column', '').strip() if form.cleaned_data.get('description_column') else None,
                    'category': form.cleaned_data.get('category_column', '').strip() if form.cleaned_data.get('category_column') else None,
                    'restaurant': form.cleaned_data.get('restaurant_column', '').strip() if form.cleaned_data.get('restaurant_column') else None,
                }
                
                try:
                    import pandas as pd
                    from dateutil import parser as date_parser
                    
                    # Читаем Excel файл
                    xls = pd.ExcelFile(excel_file)
                    
                    # Получаем или создаем ресторан по умолчанию
                    default_restaurant = None
                    if default_restaurant_name:
                        default_restaurant, _ = Restaurant.objects.get_or_create(
                            name=default_restaurant_name,
                            defaults={'is_active': True}
                        )
                    
                    # Статистика импорта
                    groups_created = 0
                    items_created = 0
                    items_updated = 0
                    errors = []
                    
                    # Обрабатываем каждый лист как отдельную группу (меню на день)
                    for sheet_name in xls.sheet_names:
                        try:
                            # Читаем без заголовков, чтобы найти строку с заголовками
                            df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                            
                            # Ищем строку с заголовками "Наименование" и "Цена"
                            header_row_idx = None
                            name_col_idx = None
                            price_col_idx = None
                            
                            for idx in range(min(10, len(df_raw))):  # Проверяем первые 10 строк
                                row_values = df_raw.iloc[idx].astype(str).str.lower().str.strip()
                                
                                # Ищем столбцы с "наименование" и "цена"
                                name_cols = [i for i, v in enumerate(row_values) if 'наименование' in str(v).lower() or 'название' in str(v).lower()]
                                price_cols = [i for i, v in enumerate(row_values) if 'цена' in str(v).lower() or 'стоимость' in str(v).lower()]
                                
                                if name_cols and price_cols:
                                    header_row_idx = idx
                                    name_col_idx = name_cols[0]
                                    price_col_idx = price_cols[0]
                                    break
                            
                            # Если не нашли заголовки автоматически, используем значения из формы или стандартные
                            if header_row_idx is None:
                                df = pd.read_excel(xls, sheet_name=sheet_name)
                                # Используем маппинг из формы или ищем автоматически
                                if column_mapping.get('name'):
                                    name_col = column_mapping['name']
                                else:
                                    name_col = next((col for col in df.columns if 'наименование' in str(col).lower() or 'название' in str(col).lower()), None)
                                
                                if column_mapping.get('price'):
                                    price_col = column_mapping['price']
                                else:
                                    price_col = next((col for col in df.columns if 'цена' in str(col).lower() or 'стоимость' in str(col).lower()), None)
                            else:
                                # Используем найденную строку как заголовки
                                df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row_idx)
                                # Обновляем маппинг на основе найденных столбцов
                                name_col = df.columns[name_col_idx] if name_col_idx < len(df.columns) else None
                                price_col = df.columns[price_col_idx] if price_col_idx < len(df.columns) else None
                                
                                # Если не нашли, пробуем найти по названию
                                if not name_col:
                                    name_col = next((col for col in df.columns if 'наименование' in str(col).lower() or 'название' in str(col).lower()), None)
                                if not price_col:
                                    price_col = next((col for col in df.columns if 'цена' in str(col).lower() or 'стоимость' in str(col).lower()), None)
                                
                                # Если указаны в форме, используем их
                                if column_mapping.get('name') and column_mapping['name'] in df.columns:
                                    name_col = column_mapping['name']
                                if column_mapping.get('price') and column_mapping['price'] in df.columns:
                                    price_col = column_mapping['price']
                            
                            # Пытаемся определить дату из названия листа
                            group_date = None
                            try:
                                # Формат: "01_12_2025" -> "01.12.2025"
                                if '_' in sheet_name:
                                    parts = sheet_name.split('_')
                                    if len(parts) == 3:
                                        date_str = f"{parts[0]}.{parts[1]}.{parts[2]}"
                                        group_date = date_parser.parse(date_str, dayfirst=True).date()
                                else:
                                    group_date = date_parser.parse(sheet_name, dayfirst=True).date()
                            except:
                                pass
                            
                            # Создаем или получаем группу
                            group_name = f"Меню на {group_date.strftime('%d.%m.%Y')}" if group_date else sheet_name
                            group, group_created = MenuItemGroup.objects.get_or_create(
                                name=group_name,
                                defaults={
                                    'is_active': True,
                                    'period_start': group_date,
                                    'period_end': group_date,
                                }
                            )
                            
                            if group_created:
                                groups_created += 1
                            
                            # Обрабатываем каждую строку
                            for idx, row in df.iterrows():
                                try:
                                    # Пропускаем пустые строки и строки-категории
                                    if pd.isna(row.get(name_col)) if name_col else True:
                                        continue
                                    
                                    name = str(row[name_col]).strip()
                                    if not name or name == 'nan' or len(name) < 2:
                                        continue
                                    
                                    # Пропускаем строки, которые выглядят как категории (короткие названия без цен)
                                    if name.isupper() or (len(name) < 15 and pd.isna(row.get(price_col)) if price_col else True):
                                        continue
                                    
                                    # Получаем цену
                                    price = None
                                    if price_col and price_col in df.columns:
                                        price_val = row[price_col]
                                        
                                        # Если это число, используем его
                                        if pd.notna(price_val):
                                            try:
                                                if isinstance(price_val, (int, float)):
                                                    price = Decimal(str(price_val))
                                                else:
                                                    price_str = str(price_val).strip()
                                                    # Очищаем от лишних символов
                                                    price_str = price_str.replace(' ', '').replace(',', '.').replace('руб', '').replace('₽', '').replace('р.', '').replace('р', '')
                                                    # Удаляем все нецифровые символы кроме точки
                                                    price_str = ''.join(c for c in price_str if c.isdigit() or c == '.')
                                                    if price_str:
                                                        price = Decimal(price_str)
                                            except (InvalidOperation, ValueError, TypeError) as e:
                                                pass
                                    
                                    # Получаем описание
                                    description = ''
                                    if column_mapping.get('description'):
                                        desc_col = column_mapping['description']
                                        if desc_col in df.columns:
                                            desc_val = row[desc_col]
                                            if pd.notna(desc_val):
                                                description = str(desc_val).strip()

                                    # Получаем категорию
                                    category = None
                                    if column_mapping.get('category'):
                                        cat_col = column_mapping['category']
                                        if cat_col in df.columns:
                                            cat_val = row[cat_col]
                                            if pd.notna(cat_val) and str(cat_val).strip():
                                                cat_name = str(cat_val).strip()
                                                category, _ = ProductCategory.objects.get_or_create(name=cat_name)
                                    
                                    # Получаем ресторан
                                    restaurant = default_restaurant
                                    if column_mapping.get('restaurant'):
                                        rest_col = column_mapping['restaurant']
                                        if rest_col in df.columns:
                                            rest_val = row[rest_col]
                                            if pd.notna(rest_val) and str(rest_val).strip():
                                                rest_name = str(rest_val).strip()
                                                restaurant, _ = Restaurant.objects.get_or_create(
                                                    name=rest_name,
                                                    defaults={'is_active': True}
                                                )
                                    
                                    if not restaurant:
                                        errors.append(f'Лист "{sheet_name}", строка {idx+2}: Не указан ресторан для товара "{name}"')
                                        continue
                                    
                                    # Создаем или обновляем товар
                                    item, item_created = MenuItem.objects.update_or_create(
                                        name=name,
                                        restaurant=restaurant,
                                        group=group,
                                        defaults={
                                            'price': price,
                                            'description': description,
                                            'category': category,
                                            'is_available': True,
                                        }
                                    )
                                    
                                    if item_created:
                                        items_created += 1
                                    else:
                                        items_updated += 1
                                        
                                except Exception as e:
                                    errors.append(f'Лист "{sheet_name}", строка {idx+2}: {str(e)}')
                            
                        except Exception as e:
                            errors.append(f'Ошибка при обработке листа "{sheet_name}": {str(e)}')
                    
                    # Формируем сообщение о результатах
                    message_parts = []
                    if groups_created:
                        message_parts.append(f'Создано групп: {groups_created}')
                    if items_created:
                        message_parts.append(f'Создано товаров: {items_created}')
                    if items_updated:
                        message_parts.append(f'Обновлено товаров: {items_updated}')
                    if errors:
                        message_parts.append(f'Ошибок: {len(errors)}')
                    
                    message = 'Импорт завершен. ' + ', '.join(message_parts)
                    
                    if errors:
                        self.message_user(request, message + '\n\nПервые ошибки:\n' + '\n'.join(errors[:10]), level='warning')
                    else:
                        self.message_user(request, message, level='success')
                    
                    return redirect('..')
                    
                except ImportError:
                    self.message_user(request, 'Ошибка: не установлены библиотеки pandas и openpyxl. Установите их через pip install pandas openpyxl', level='error')
                except Exception as e:
                    self.message_user(request, f'Ошибка при чтении файла: {str(e)}', level='error')
        else:
            form = ExcelImportForm()
        
        context = {
            'form': form,
            'title': 'Импорт групп товаров из Excel',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/orders/menuitemgroup/import_excel.html', context)
    
    def export_template(self, request):
        """Экспорт шаблона Excel для создания групп товаров"""
        from django.http import HttpResponse
        import pandas as pd
        from io import BytesIO
        from datetime import datetime, timedelta
        
        # Создаем Excel файл в памяти
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Создаем примеры листов для недели
            for i in range(7):
                date = datetime.now().date() + timedelta(days=i)
                sheet_name = date.strftime('%d.%m.%Y')
                
                # Создаем DataFrame с примерными данными
                sample_data = {
                    'Название': ['Борщ', 'Салат Цезарь', 'Пицца Маргарита'],
                    'Цена': [150, 200, 350],
                    'Категория': ['Супы', 'Салаты', 'Пицца'],
                    'Ресторан': ['ВкусВилл', 'ВкусВилл', 'Додо Пицца'],
                    'Описание': ['Традиционный борщ', 'С курицей и соусом', 'Классическая пицца']
                }
                df = pd.DataFrame(sample_data)
                
                # Записываем в лист
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="menu_template.xlsx"'
        return response
    
    def export_orders_report(self, request, queryset):
        """Экспорт отчёта по заказам для выбранных групп (Excel и CSV)"""
        from django.http import HttpResponse
        import pandas as pd
        from io import BytesIO
        from django.utils import timezone
        
        # Получаем формат экспорта из параметров
        export_format = request.GET.get('format', 'excel')
        
        # Получаем все заказы для товаров из выбранных групп
        from .models import OrderItem
        order_items = OrderItem.objects.filter(
            menu_item__group__in=queryset
        ).select_related('order', 'order__employee', 'menu_item', 'menu_item__group')
        
        # Формируем данные
        data = []
        for item in order_items:
            data.append({
                'Дата заказа': item.order.order_date.strftime('%Y-%m-%d'),
                'Сотрудник': item.order.employee.name,
                'Email': item.order.employee.email,
                'Telegram ID': item.order.employee.telegram_id or '',
                'Группа товаров': item.menu_item.group.name if item.menu_item.group else '',
                'Блюдо': item.menu_item.name,
                'Количество': item.quantity,
                'Цена': float(item.price),
                'Сумма': float(item.subtotal),
                'Статус': item.order.get_status_display(),
            })
        
        df = pd.DataFrame(data)
        
        if export_format == 'excel':
            # Экспорт в Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Заказы', index=False)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'
        else:
            # Экспорт в CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="orders_report.csv"'
            response.write(df.to_csv(index=False, encoding='utf-8-sig'))
        
        return response
    export_orders_report.short_description = 'Экспорт отчёта по заказам (Excel/CSV)'


class MenuItemImageInline(admin.TabularInline):
    model = MenuItemImage
    extra = 1
    fields = ['image', 'is_primary', 'order']


@admin.register(ProductCategory)
class ProductCategoryAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['name', 'order', 'is_active']
    list_editable = ['order', 'is_active']
    search_fields = ['name']


@admin.register(MenuItem)
class MenuItemAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['name', 'group', 'category', 'restaurant', 'price', 'is_available', 'image_count']
    list_filter = ['restaurant', 'group', 'category', 'is_available']
    search_fields = ['name', 'description']
    inlines = [MenuItemImageInline]
    change_list_template = 'admin/orders/menuitem/change_list.html'
    
    def image_count(self, obj):
        return obj.images.count()
    image_count.short_description = 'Изображений'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.admin_site.admin_view(self.import_csv), name='orders_menuitem_import_csv'),
            path('upload-image-links/', self.admin_site.admin_view(self.upload_image_links), name='orders_menuitem_upload_image_links'),
        ]
        return custom_urls + urls

    def upload_image_links(self, request):
        """Загрузка image_links.csv: по названию/группе/ресторану находим блюдо и ставим image_url (показывается, если нет загруженной картинки)."""
        if request.method == 'POST' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            try:
                content = csv_file.read()
                for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
                    try:
                        text = content.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    self.message_user(request, 'Не удалось определить кодировку CSV. Используйте UTF-8.', level='error')
                    return redirect('admin:orders_menuitem_upload_image_links')
                # Разделитель: в файле может быть запятая или точка с запятой
                for delim in (',', ';'):
                    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
                    rows = list(reader)
                    if rows and (rows[0].get('Название') or rows[0].get('Изображение')):
                        break
                else:
                    rows = []
                updated = 0
                not_found = []
                for row in rows:
                    name = (row.get('Название') or '').strip()
                    img_url = (row.get('Изображение') or '').strip()
                    if not name or not img_url:
                        continue
                    rest_name = (row.get('Ресторан') or '').strip()
                    qs = MenuItem.objects.filter(name=name)
                    if rest_name:
                        qs = qs.filter(restaurant__name=rest_name)
                    items = list(qs)
                    if items:
                        for item in items:
                            item.image_url = img_url
                            item.save()
                            updated += 1
                    else:
                        not_found.append(name)
                msg = f'Обновлено ссылок на изображения: {updated}.'
                if not_found:
                    msg += f' Не найдено блюд: {", ".join(not_found[:10])}{"…" if len(not_found) > 10 else ""}'
                self.message_user(request, msg, level='success')
            except Exception as e:
                self.message_user(request, f'Ошибка: {str(e)}', level='error')
            return redirect('admin:orders_menuitem_changelist')
        context = {
            'title': 'Загрузить image_links.csv',
            'opts': self.model._meta,
        }
        return render(request, 'admin/orders/menuitem/upload_image_links.html', context)
    
    def import_csv(self, request):
        """Импорт товаров из CSV с поддержкой ZIP архива с изображениями"""
        if request.method == 'POST':
            form = CSVImportForm(request.POST, request.FILES)
            if form.is_valid():
                csv_file = request.FILES['csv_file']
                zip_file = request.FILES.get('zip_file')
                encoding = form.cleaned_data['encoding']
                delimiter = form.cleaned_data['delimiter']
                
                # Маппинг столбцов
                column_mapping = {
                    'name': form.cleaned_data['name_column'].lower().strip(),
                    'group': form.cleaned_data['group_column'].lower().strip(),
                    'restaurant': form.cleaned_data['restaurant_column'].lower().strip(),
                    'price': form.cleaned_data.get('price_column', '').lower().strip(),
                    'description': form.cleaned_data.get('description_column', '').lower().strip(),
                    'image': form.cleaned_data.get('image_column', '').lower().strip(),
                }
                
                # Распаковка ZIP архива если он есть
                zip_extract_path = None
                zip_file_dict = {}
                if zip_file:
                    try:
                        zip_extract_path = tempfile.mkdtemp()
                        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                            zip_ref.extractall(zip_extract_path)
                            # Создаем словарь для быстрого поиска файлов
                            for root, dirs, files in os.walk(zip_extract_path):
                                for file in files:
                                    rel_path = os.path.relpath(os.path.join(root, file), zip_extract_path)
                                    # Нормализуем путь (заменяем обратные слеши на прямые)
                                    rel_path = rel_path.replace('\\', '/')
                                    zip_file_dict[rel_path] = os.path.join(root, file)
                                    # Также добавляем вариант без учета регистра
                                    zip_file_dict[rel_path.lower()] = os.path.join(root, file)
                    except Exception as e:
                        self.message_user(request, f'Ошибка при распаковке ZIP архива: {str(e)}', level='error')
                        form = CSVImportForm()
                        context = {
                            'form': form,
                            'title': 'Импорт товаров из CSV',
                            'opts': self.model._meta,
                            'has_view_permission': self.has_view_permission(request),
                        }
                        return render(request, 'admin/orders/menuitem/import_csv.html', context)
                
                try:
                    # Читаем CSV
                    decoded_file = csv_file.read().decode(encoding)
                    io_string = io.StringIO(decoded_file)
                    reader = csv.DictReader(io_string, delimiter=delimiter)
                    
                    # Нормализуем заголовки (приводим к нижнему регистру)
                    normalized_reader = []
                    for row in reader:
                        normalized_row = {k.lower().strip(): v for k, v in row.items()}
                        normalized_reader.append(normalized_row)
                    
                    # Статистика импорта
                    created_count = 0
                    updated_count = 0
                    error_count = 0
                    image_count = 0
                    errors = []
                    
                    for row_num, row in enumerate(normalized_reader, start=2):
                        try:
                            # Получаем значения из CSV
                            name = row.get(column_mapping['name'], '').strip()
                            group_name = row.get(column_mapping['group'], '').strip()
                            restaurant_name = row.get(column_mapping['restaurant'], '').strip()
                            price_str = row.get(column_mapping['price'], '').strip() if column_mapping['price'] else ''
                            description = row.get(column_mapping['description'], '').strip() if column_mapping['description'] else ''
                            image_path = row.get(column_mapping['image'], '').strip() if column_mapping['image'] else ''
                            
                            # Валидация обязательных полей
                            if not name:
                                errors.append(f'Строка {row_num}: Отсутствует название товара')
                                error_count += 1
                                continue
                            
                            if not group_name:
                                errors.append(f'Строка {row_num}: Отсутствует группа товара')
                                error_count += 1
                                continue
                            
                            if not restaurant_name:
                                errors.append(f'Строка {row_num}: Отсутствует ресторан')
                                error_count += 1
                                continue
                            
                            # Получаем или создаем группу
                            group, _ = MenuItemGroup.objects.get_or_create(
                                name=group_name,
                                defaults={'is_active': True}
                            )
                            
                            # Получаем или создаем ресторан
                            restaurant, _ = Restaurant.objects.get_or_create(
                                name=restaurant_name,
                                defaults={'is_active': True}
                            )
                            
                            # Парсим цену
                            price = None
                            if price_str:
                                try:
                                    # Убираем пробелы и заменяем запятую на точку
                                    price_str = price_str.replace(' ', '').replace(',', '.')
                                    price = Decimal(price_str)
                                except (InvalidOperation, ValueError):
                                    errors.append(f'Строка {row_num}: Некорректная цена "{price_str}"')
                            
                            # Создаем или обновляем товар
                            item, created = MenuItem.objects.update_or_create(
                                name=name,
                                restaurant=restaurant,
                                defaults={
                                    'group': group,
                                    'price': price,
                                    'description': description,
                                    'is_available': True,
                                }
                            )
                            
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                            
                            # Обработка изображения из ZIP архива
                            if image_path and zip_file and zip_extract_path:
                                try:
                                    # Нормализуем путь (заменяем обратные слеши на прямые)
                                    image_path = image_path.replace('\\', '/')
                                    
                                    # Ищем файл в распакованном архиве
                                    image_file_path = None
                                    if image_path in zip_file_dict:
                                        image_file_path = zip_file_dict[image_path]
                                    elif image_path.lower() in zip_file_dict:
                                        image_file_path = zip_file_dict[image_path.lower()]
                                    else:
                                        # Пробуем найти файл по имени без учета регистра
                                        image_name = os.path.basename(image_path)
                                        for rel_path, full_path in zip_file_dict.items():
                                            if os.path.basename(rel_path).lower() == image_name.lower():
                                                image_file_path = full_path
                                                break
                                    
                                    if image_file_path and os.path.exists(image_file_path):
                                        # Создаем MenuItemImage
                                        with open(image_file_path, 'rb') as img_file:
                                            django_file = File(img_file, name=os.path.basename(image_path))
                                            menu_item_image = MenuItemImage.objects.create(
                                                menu_item=item,
                                                image=django_file,
                                                is_primary=(item.images.count() == 0),  # Первое изображение - основное
                                                order=item.images.count()
                                            )
                                            image_count += 1
                                    else:
                                        errors.append(f'Строка {row_num}: Изображение не найдено в архиве: {image_path}')
                                except Exception as e:
                                    errors.append(f'Строка {row_num}: Ошибка при загрузке изображения "{image_path}": {str(e)}')
                                
                        except Exception as e:
                            errors.append(f'Строка {row_num}: {str(e)}')
                            error_count += 1
                    
                    # Удаляем временную директорию с распакованным архивом
                    if zip_extract_path and os.path.exists(zip_extract_path):
                        import shutil
                        try:
                            shutil.rmtree(zip_extract_path)
                        except:
                            pass
                    
                    # Формируем сообщение о результатах
                    message_parts = []
                    if created_count:
                        message_parts.append(f'Создано: {created_count}')
                    if updated_count:
                        message_parts.append(f'Обновлено: {updated_count}')
                    if image_count:
                        message_parts.append(f'Загружено изображений: {image_count}')
                    if error_count:
                        message_parts.append(f'Ошибок: {error_count}')
                    
                    message = 'Импорт завершен. ' + ', '.join(message_parts)
                    
                    if errors:
                        self.message_user(request, message + '\n\nОшибки:\n' + '\n'.join(errors[:10]), level='warning')
                    else:
                        self.message_user(request, message, level='success')
                    
                    return redirect('..')
                    
                except Exception as e:
                    # Удаляем временную директорию в случае ошибки
                    if zip_extract_path and os.path.exists(zip_extract_path):
                        import shutil
                        try:
                            shutil.rmtree(zip_extract_path)
                        except:
                            pass
                    self.message_user(request, f'Ошибка при чтении файла: {str(e)}', level='error')
        else:
            form = CSVImportForm()
        
        context = {
            'form': form,
            'title': 'Импорт товаров из CSV',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/orders/menuitem/import_csv.html', context)


@admin.register(MenuItemImage)
class MenuItemImageAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['menu_item', 'is_primary', 'order', 'created_at']
    list_filter = ['is_primary', 'created_at']
    search_fields = ['menu_item__name']


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0


@admin.register(Order)
class OrderAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['id', 'employee', 'group', 'order_date', 'total_amount', 'status']
    list_filter = ['status', 'order_date', 'group']
    search_fields = ['employee__name', 'employee__email', 'group__name']
    inlines = [OrderItemInline]
    actions = ['export_all_orders_report']
    
    def export_all_orders_report(self, request, queryset):
        """Экспорт единого отчёта по всем заказам (Excel и CSV)"""
        from django.http import HttpResponse
        import pandas as pd
        from io import BytesIO
        
        # Получаем формат экспорта из параметров
        export_format = request.GET.get('format', 'excel')
        
        # Если выбраны конкретные заказы, используем их, иначе все
        if queryset.exists():
            orders = queryset
        else:
            orders = Order.objects.all()
        
        # Получаем все позиции заказов
        from .models import OrderItem
        order_items = OrderItem.objects.filter(
            order__in=orders
        ).select_related('order', 'order__employee', 'menu_item', 'menu_item__group', 'menu_item__restaurant')
        
        # Формируем данные
        data = []
        for item in order_items:
            data.append({
                'Дата заказа': item.order.order_date.strftime('%Y-%m-%d'),
                'Сотрудник': item.order.employee.name,
                'Email': item.order.employee.email,
                'Telegram ID': item.order.employee.telegram_id or '',
                'Группа товаров': item.menu_item.group.name if item.menu_item.group else '',
                'Блюдо': item.menu_item.name,
                'Ресторан': item.menu_item.restaurant.name,
                'Количество': item.quantity,
                'Цена': float(item.price),
                'Сумма': float(item.subtotal),
                'Статус заказа': item.order.get_status_display(),
            })
        
        df = pd.DataFrame(data)
        
        if export_format == 'excel':
            # Экспорт в Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Все заказы', index=False)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="all_orders_report.xlsx"'
        else:
            # Экспорт в CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="all_orders_report.csv"'
            response.write(df.to_csv(index=False, encoding='utf-8-sig'))
        
        return response
    export_all_orders_report.short_description = 'Экспорт единого отчёта по всем заказам (Excel/CSV)'
    
    change_list_template = 'admin/orders/order/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-csv/', self.admin_site.admin_view(self.export_csv_view), name='orders_order_export_csv'),
            path('export-xlsx/', self.admin_site.admin_view(self.export_xlsx_view), name='orders_order_export_xlsx'),
        ]
        return custom_urls + urls

    def export_csv_view(self, request):
        if request.method == 'POST':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            if start_date and end_date:
                # Generate CSV
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="orders_{start_date}_{end_date}.csv"'
                response.write('\ufeff')  # BOM for Excel UTF-8
                
                writer = csv.writer(response)
                writer.writerow(['ID', 'Дата', 'Сотрудник', 'Telegram ID', 'Блюдо', 'Категория', 'Количество', 'Цена', 'Сумма', 'Статус'])
                
                from .models import OrderItem
                items = OrderItem.objects.filter(
                    order__order_date__range=[start_date, end_date]
                ).select_related('order', 'order__employee', 'menu_item', 'menu_item__category')
                
                for item in items:
                    writer.writerow([
                        item.order.id,
                        item.order.order_date,
                        item.order.employee.name,
                        item.order.employee.telegram_id,
                        item.menu_item.name,
                        item.menu_item.category.name if item.menu_item.category else 'Остальное',
                        item.quantity,
                        item.price,
                        item.subtotal,
                        item.order.get_status_display()
                    ])
                
                return response
        
        context = {
            'title': 'Экспорт заказов в CSV',
            'opts': self.model._meta,
            'today': datetime.now().strftime('%Y-%m-%d'),
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/orders/order/export_csv.html', context)

    def export_xlsx_view(self, request):
        if request.method == 'POST':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            if start_date and end_date:
                import pandas as pd
                from io import BytesIO
                
                # Generate XLSX
                output = BytesIO()
                
                from .models import OrderItem
                items = OrderItem.objects.filter(
                    order__order_date__range=[start_date, end_date]
                ).select_related('order', 'order__employee', 'menu_item', 'menu_item__category')
                
                data = []
                for item in items:
                    data.append({
                        'ID': item.order.id,
                        'Дата': item.order.order_date,
                        'Сотрудник': item.order.employee.name,
                        'Telegram ID': item.order.employee.telegram_id,
                        'Блюдо': item.menu_item.name,
                        'Категория': item.menu_item.category.name if item.menu_item.category else 'Остальное',
                        'Количество': item.quantity,
                        'Цена': float(item.price),
                        'Сумма': float(item.subtotal),
                        'Статус': item.order.get_status_display()
                    })
                
                df = pd.DataFrame(data)
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Заказы', index=False)
                
                output.seek(0)
                
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="orders_{start_date}_{end_date}.xlsx"'
                return response
        
        context = {
            'title': 'Экспорт заказов в XLSX',
            'opts': self.model._meta,
            'today': datetime.now().strftime('%Y-%m-%d'),
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/orders/order/export_xlsx.html', context)


@admin.register(BalanceTransaction)
class BalanceTransactionAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['employee', 'transaction_type', 'amount', 'balance_after', 'created_at', 'order']
    list_filter = ['transaction_type', 'created_at']
    search_fields = ['employee__name', 'employee__email', 'comment']
    readonly_fields = ['created_at', 'balance_after']
    date_hierarchy = 'created_at'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('employee', 'transaction_type', 'amount', 'balance_after')
        }),
        ('Дополнительно', {
            'fields': ('order', 'comment', 'created_at')
        }),
    )

    def save_model(self, request, obj, form, change):
        if not change:  # Only on creation
            if obj.transaction_type == 'accrual':
                obj.employee.balance += obj.amount
            elif obj.transaction_type == 'deduction':
                obj.employee.balance -= obj.amount
            elif obj.transaction_type == 'correction':
                obj.employee.balance += obj.amount
            elif obj.transaction_type == 'refund':
                obj.employee.balance += obj.amount
            
            obj.balance_after = obj.employee.balance
            obj.employee.save()
        
        super().save_model(request, obj, form, change)


@admin.register(WorkDayCalendar)
class WorkDayCalendarAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['employee', 'date', 'day_type', 'comment', 'created_at']
    list_filter = ['day_type', 'date']
    search_fields = ['employee__name', 'employee__email', 'comment']
    date_hierarchy = 'date'
    readonly_fields = ['created_at', 'updated_at']
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('employee', 'date', 'day_type')
        }),
        ('Дополнительно', {
            'fields': ('comment', 'created_at', 'updated_at')
        }),
    )


@admin.register(Settings)
class SettingsAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ['key', 'value', 'description']


@admin.register(SystemConfig)
class SystemConfigAdmin(ExportCsvMixin, admin.ModelAdmin):
    actions = ['export_as_csv']
    list_display = ('__str__', 'logo')
    
    def has_add_permission(self, request):
        # Allow adding only if no instance exists
        if self.model.objects.exists():
            return False
        return super().has_add_permission(request)


@admin.register(GlobalWorkDay)
class GlobalWorkDayAdmin(ExportCsvMixin, admin.ModelAdmin):
    """Глобальный календарь рабочих дней для всех сотрудников"""
    actions = ['export_as_csv']
    list_display = ['date', 'day_type', 'comment', 'created_at']
    list_filter = ['day_type']
    search_fields = ['comment']
    date_hierarchy = 'date'
    ordering = ['-date']
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('date', 'day_type', 'comment')
        }),
    )


# ============== CUSTOM PRODUCT IMPORT VIEW ==============

class ProductImportForm(forms.Form):
    """Форма для импорта товаров из Excel"""
    excel_file = forms.FileField(label='Excel файл (.xlsx)')


def product_import_view(request):
    """Страница импорта товаров с шаблоном"""
    from django.contrib.admin.views.decorators import staff_member_required
    from openpyxl import Workbook, load_workbook
    
    context = {
        'title': 'Импорт товаров из Excel',
        'has_permission': True,
    }
    
    if request.method == 'POST':
        if 'download_template' in request.POST:
            # Скачать шаблон Excel
            wb = Workbook()
            ws = wb.active
            ws.title = 'Товары'
            
            # Заголовки
            headers = ['Название', 'Описание', 'Цена', 'Категория (обязательно)', 'Группа', 'Ресторан', 'Доступен (да/нет)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Примеры данных
            example_data = [
                ['Борщ украинский', 'Классический борщ со сметаной', 250, 'Первые блюда', 'Обед', 'ВкусВилл', 'да'],
                ['Котлета по-киевски', 'Куриная котлета с маслом', 350, 'Вторые блюда', 'Обед', 'ВкусВилл', 'да'],
                ['Салат Цезарь', 'С курицей и соусом', 280, 'Салаты', 'Обед', 'ВкусВилл', 'да'],
            ]
            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Добавляем лист с категориями
            ws_cat = wb.create_sheet('Категории (справочник)')
            ws_cat.cell(row=1, column=1, value='Доступные категории:')
            categories = ProductCategory.objects.filter(is_active=True).order_by('order')
            for idx, cat in enumerate(categories, 2):
                ws_cat.cell(row=idx, column=1, value=cat.name)
            
            # Добавляем лист с группами
            ws_grp = wb.create_sheet('Группы (справочник)')
            ws_grp.cell(row=1, column=1, value='Доступные группы:')
            groups = MenuItemGroup.objects.filter(is_active=True).order_by('order')
            for idx, grp in enumerate(groups, 2):
                ws_grp.cell(row=idx, column=1, value=grp.name)
            
            # Добавляем лист с ресторанами
            ws_rest = wb.create_sheet('Рестораны (справочник)')
            ws_rest.cell(row=1, column=1, value='Доступные рестораны:')
            restaurants = Restaurant.objects.filter(is_active=True)
            for idx, rest in enumerate(restaurants, 2):
                ws_rest.cell(row=idx, column=1, value=rest.name)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
            wb.save(response)
            return response
        
        elif 'import_file' in request.POST:
            # Импорт из Excel
            form = ProductImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    
                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    imported = 0
                    errors = []
                    
                    for row_idx, row in enumerate(rows, 2):
                        if not row or not row[0]:  # Пропускаем пустые строки
                            continue
                        
                        name = str(row[0]).strip() if row[0] else ''
                        description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                        price_str = str(row[2]).strip() if len(row) > 2 and row[2] else '0'
                        category_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                        group_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                        restaurant_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        is_available_str = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'да'
                        
                        if not name:
                            continue
                        
                        # Проверка обязательного поля - категория
                        if not category_name:
                            errors.append(f'Строка {row_idx}: Не указана категория для "{name}"')
                            continue
                        
                        # Находим или создаём категорию
                        category, _ = ProductCategory.objects.get_or_create(
                            name=category_name,
                            defaults={'is_active': True}
                        )
                        
                        # Находим группу
                        group = None
                        if group_name:
                            group = MenuItemGroup.objects.filter(name__iexact=group_name).first()
                        
                        # Находим ресторан
                        restaurant = None
                        if restaurant_name:
                            restaurant = Restaurant.objects.filter(name__iexact=restaurant_name).first()
                        if not restaurant:
                            restaurant = Restaurant.objects.first()
                        
                        # Парсим цену
                        try:
                            price = Decimal(str(price_str).replace(',', '.').replace(' ', ''))
                        except:
                            price = Decimal('0')
                        
                        is_available = is_available_str in ['да', 'yes', '1', 'true']
                        
                        # Создаём или обновляем товар
                        menu_item, created = MenuItem.objects.update_or_create(
                            name=name,
                            restaurant=restaurant,
                            defaults={
                                'description': description,
                                'price': price,
                                'category': category,
                                'group': group,
                                'is_available': is_available,
                            }
                        )
                        imported += 1
                    
                    context['success'] = f'Успешно импортировано: {imported} товаров'
                    if errors:
                        context['errors'] = errors
                        
                except Exception as e:
                    context['error'] = f'Ошибка при импорте: {str(e)}'
            else:
                context['error'] = 'Выберите файл для импорта'
    
    context['form'] = ProductImportForm()
    return render(request, 'admin/orders/product_import.html', context)


# Регистрируем URL в admin
from django.urls import path

class CustomAdminSite(admin.AdminSite):
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('product-import/', self.admin_view(product_import_view), name='product_import'),
        ]
        return custom_urls + urls

# Добавляем URL напрямую к стандартному admin
original_get_urls = admin.site.get_urls
def custom_get_urls():
    urls = original_get_urls()
    custom_urls = [
        path('product-import/', admin.site.admin_view(product_import_view), name='product_import'),
    ]
    return custom_urls + urls

admin.site.get_urls = custom_get_urls


