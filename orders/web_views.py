from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from .models import MenuItem, MenuItemGroup, MenuItemImage, Restaurant, OrderItem, Order, Employee, WeekCompanyAmount
from decimal import Decimal
from django.db.models import Sum, F
from django.utils import timezone
import json
import csv
import os
from datetime import datetime, timedelta


def is_superuser(user):
    """Проверка, является ли пользователь суперпользователем"""
    return user.is_superuser


def login_view(request):
    """Страница входа"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('admin_products')
        else:
            return render(request, 'orders/login.html', {'error': 'Неверные учетные данные'})
    return render(request, 'orders/login.html')


def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('landing')


def landing_page(request):
    """Главная страница - лендинг о стартапе"""
    return render(request, 'orders/landing.html')


@csrf_exempt
@require_http_methods(["POST"])
def contact_form_submit(request):
    """Обработка формы обратной связи и сохранение в CSV"""
    try:
        data = json.loads(request.body)
        
        # Получаем данные из формы
        name = data.get('name', '')
        email = data.get('email', '')
        phone = data.get('phone', '')
        company = data.get('company', '')
        message = data.get('message', '')
        form_type = data.get('form_type', 'partner')  # 'partner' или 'catering'
        
        # Путь к CSV файлу в статике
        csv_file_path = os.path.join(settings.STATIC_ROOT, 'i_am_secret_file.csv')
        
        # Создаём директорию если её нет
        os.makedirs(os.path.dirname(csv_file_path), exist_ok=True)
        
        # Проверяем, существует ли файл, если нет - создаём с заголовками
        file_exists = os.path.isfile(csv_file_path)
        
        # Записываем данные в CSV
        with open(csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['timestamp', 'form_type', 'name', 'email', 'phone', 'company', 'message']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            if not file_exists:
                writer.writeheader()
            
            writer.writerow({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'form_type': form_type,
                'name': name,
                'email': email,
                'phone': phone,
                'company': company,
                'message': message
            })
        
        return JsonResponse({'success': True, 'message': 'Спасибо! Мы свяжемся с вами в ближайшее время.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_csv_file(request):
    """Доступ к CSV файлу с заявками"""
    csv_file_path = os.path.join(settings.STATIC_ROOT, 'i_am_secret_file.csv')
    
    if os.path.exists(csv_file_path):
        response = FileResponse(open(csv_file_path, 'rb'), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="i_am_secret_file.csv"'
        return response
    else:
        return HttpResponse('File not found', status=404)


def product_catalog(request):
    """Публичный каталог товаров (для незалогиненных пользователей и mini-app)"""
    # Получаем только доступные товары из активных ресторанов
    from django.utils import timezone
    today = timezone.now().date()
    
    # Фильтруем активные рестораны с учетом периодов
    active_restaurants = Restaurant.objects.filter(
        Q(is_active=True) &
        (Q(period_start__isnull=True) | Q(period_start__lte=today)) &
        (Q(period_end__isnull=True) | Q(period_end__gte=today))
    )
    
    # Получаем доступные товары
    items = MenuItem.objects.filter(
        restaurant__in=active_restaurants,
        is_available=True,
        price__isnull=False
    ).select_related('restaurant', 'group').prefetch_related('images')
    
    # Группируем по категориям
    groups = MenuItemGroup.objects.filter(
        is_active=True,
        items__in=items
    ).distinct().prefetch_related('items')
    
    context = {
        'groups': groups,
        'items': items,
        'is_authenticated': request.user.is_authenticated,
    }
    
    return render(request, 'orders/catalog.html', context)


@login_required
@user_passes_test(is_superuser)
def admin_products(request):
    """Административная панель управления товарами"""
    items = MenuItem.objects.all().select_related('restaurant', 'group').prefetch_related('images')
    groups = MenuItemGroup.objects.all()
    restaurants = Restaurant.objects.all()
    
    context = {
        'items': items,
        'groups': groups,
        'restaurants': restaurants,
    }
    
    return render(request, 'orders/admin_products.html', context)


@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def create_product(request):
    """Создание нового товара"""
    try:
        name = request.POST.get('name')
        group_id = request.POST.get('group')
        restaurant_id = request.POST.get('restaurant')
        description = request.POST.get('description', '')
        price = request.POST.get('price')
        
        # Валидация обязательных полей
        if not name or not group_id:
            return JsonResponse({'success': False, 'error': 'Название и группа обязательны'}, status=400)
        
        # Создаем товар
        item = MenuItem.objects.create(
            name=name,
            group_id=group_id,
            restaurant_id=restaurant_id,
            description=description,
            price=Decimal(price) if price else None,
            is_available=True
        )
        
        # Обрабатываем изображения
        images = request.FILES.getlist('images')
        for idx, image in enumerate(images):
            MenuItemImage.objects.create(
                menu_item=item,
                image=image,
                is_primary=(idx == 0),
                order=idx
            )
        
        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'name': item.name,
                'price': str(item.price) if item.price else None,
                'group': item.group.name if item.group else None,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def update_product(request, product_id):
    """Обновление товара"""
    try:
        item = get_object_or_404(MenuItem, id=product_id)
        
        item.name = request.POST.get('name', item.name)
        item.description = request.POST.get('description', item.description)
        
        group_id = request.POST.get('group')
        if group_id:
            item.group_id = group_id
        
        restaurant_id = request.POST.get('restaurant')
        if restaurant_id:
            item.restaurant_id = restaurant_id
        
        price = request.POST.get('price')
        if price:
            item.price = Decimal(price)
        
        item.save()
        
        # Обрабатываем новые изображения
        images = request.FILES.getlist('images')
        if images:
            # Получаем текущее количество изображений
            current_count = item.images.count()
            for idx, image in enumerate(images):
                MenuItemImage.objects.create(
                    menu_item=item,
                    image=image,
                    is_primary=(current_count == 0 and idx == 0),
                    order=current_count + idx
                )
        
        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'name': item.name,
                'price': str(item.price) if item.price else None,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def delete_product(request, product_id):
    """Удаление товара"""
    try:
        item = get_object_or_404(MenuItem, id=product_id)
        item.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def delete_image(request, image_id):
    """Удаление изображения товара"""
    try:
        image = get_object_or_404(MenuItemImage, id=image_id)
        image.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def set_primary_image(request, image_id):
    """Установка основного изображения"""
    try:
        image = get_object_or_404(MenuItemImage, id=image_id)
        # Снимаем флаг со всех изображений этого товара
        MenuItemImage.objects.filter(menu_item=image.menu_item).update(is_primary=False)
        # Устанавливаем флаг для выбранного
        image.is_primary = True
        image.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_product_details(request, product_id):
    """Получение деталей товара (для модального окна)"""
    try:
        item = get_object_or_404(MenuItem, id=product_id)
        images = [{'id': img.id, 'url': img.image.url, 'is_primary': img.is_primary} 
                  for img in item.images.all()]
        
        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'price': str(item.price) if item.price else None,
                'group': {'id': item.group.id, 'name': item.group.name} if item.group else None,
                'restaurant': {'id': item.restaurant.id, 'name': item.restaurant.name},
                'images': images,
                'is_available': item.is_available,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def spora_instruction(request):
    """Страница с инструкцией по адресу /spora"""
    return render(request, 'orders/spora.html')


# Названия дней недели для заголовков таблицы (как в Excel)
WEEKDAY_NAMES_RU = [
    'понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье'
]


def _get_svodka_context(request):
    """Общая логика сводки: парсит даты, считает данные, возвращает context для шаблона."""
    today = timezone.now().date()
    # По умолчанию — текущая неделя (пн–вс)
    default_start = today - timedelta(days=today.weekday())
    default_end = default_start + timedelta(days=6)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        try:
            default_start = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            default_end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if default_start > default_end:
        default_start, default_end = default_end, default_start

    num_days = (default_end - default_start).days + 1
    dates_in_range = [default_start + timedelta(days=i) for i in range(num_days)]

    from django.db.models import F
    from collections import defaultdict

    day_totals = (
        OrderItem.objects.filter(
            order__order_date__gte=default_start,
            order__order_date__lte=default_end
        )
        .exclude(order__status='cancelled')
        .values('order__employee_id', 'order__order_date')
        .annotate(day_total=Sum(F('price') * F('quantity')))
    )
    emp_date_to_total = defaultdict(Decimal)
    for row in day_totals:
        order_date = row['order__order_date']
        if hasattr(order_date, 'date'):
            order_date = order_date.date()
        key = (row['order__employee_id'], order_date)
        emp_date_to_total[key] = row['day_total'] or Decimal('0')

    employee_ids = list(
        Order.objects.filter(
            order_date__gte=default_start,
            order_date__lte=default_end
        )
        .exclude(status='cancelled')
        .values_list('employee_id', flat=True)
        .distinct()
    )
    employees = list(
        Employee.objects.filter(id__in=employee_ids).order_by('user__last_name', 'user__first_name')
        if employee_ids else []
    )

    table_rows = []
    total_fact = Decimal('0')
    totals_by_date = {d: {'fact': Decimal('0'), 'company': Decimal('0'), 'employee': Decimal('0')} for d in dates_in_range}

    for num, emp in enumerate(employees, start=1):
        fact_total = Decimal('0')
        day_cells = []
        for d in dates_in_range:
            fact_val = emp_date_to_total.get((emp.id, d), Decimal('0'))
            fact_total += fact_val
            company_val = Decimal('0')
            employee_val = Decimal('0')
            day_cells.append({'fact': fact_val, 'company': company_val, 'employee': employee_val})
            totals_by_date[d]['fact'] += fact_val
            totals_by_date[d]['company'] += company_val
            totals_by_date[d]['employee'] += employee_val
        # Для мобильных карточек: ячейки с подписанными датами
        day_cells_with_date = [
            {'date_label': d.strftime('%d.%m'), 'day_name': WEEKDAY_NAMES_RU[d.weekday()], **day_cells[i]}
            for i, d in enumerate(dates_in_range)
        ]
        table_rows.append({
            'num': num,
            'name': emp.name,
            'fact_company_prepay': Decimal('0'),
            'fact_employee_prepay': Decimal('0'),
            'day_cells': day_cells,
            'day_cells_with_date': day_cells_with_date,
            'fact_total': fact_total,
            'balance': emp.balance,
        })
        total_fact += fact_total

    total_day_cells = [totals_by_date[d] for d in dates_in_range]
    date_headers = [(d, WEEKDAY_NAMES_RU[d.weekday()]) for d in dates_in_range]
    title_colspan = 4 + 3 * len(dates_in_range) + 2

    return {
        'start_date': default_start.strftime('%Y-%m-%d'),
        'end_date': default_end.strftime('%Y-%m-%d'),
        'dates': dates_in_range,
        'date_headers': date_headers,
        'table_rows': table_rows,
        'total_day_cells': total_day_cells,
        'grand_total_fact': total_fact,
        'title_colspan': title_colspan,
    }


def _get_svodka_mob2_context(request):
    """Контекст для svodka-mob2: по неделям, компания/сотрудник/факт по дням, «сколько дала компания», остаток с прошлой недели."""
    from collections import defaultdict

    today = timezone.now().date()
    default_start = today - timedelta(days=today.weekday())
    default_end = default_start + timedelta(days=6)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    if start_date_str:
        try:
            default_start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            default_end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if default_start > default_end:
        default_start, default_end = default_end, default_start

    start = default_start
    end = default_end
    num_days = (end - start).days + 1
    dates_in_range = [start + timedelta(days=i) for i in range(num_days)]

    # «Не одна неделя» — если не пн–вс или не 7 дней
    is_one_week = (num_days == 7 and start.weekday() == 0)

    # Факт по (сотрудник, дата) из OrderItem
    day_totals = (
        OrderItem.objects.filter(
            order__order_date__gte=start,
            order__order_date__lte=end
        )
        .exclude(order__status='cancelled')
        .values('order__employee_id', 'order__order_date')
        .annotate(day_total=Sum(F('price') * F('quantity')))
    )
    emp_date_to_fact = defaultdict(lambda: Decimal('0'))
    for row in day_totals:
        order_date = row['order__order_date']
        if hasattr(order_date, 'date'):
            order_date = order_date.date()
        key = (row['order__employee_id'], order_date)
        emp_date_to_fact[key] = row['day_total'] or Decimal('0')

    # Сколько доплатил сотрудник по (сотрудник, дата) из Order.amount_paid_by_employee
    order_paid = (
        Order.objects.filter(
            order_date__gte=start,
            order_date__lte=end
        )
        .exclude(status='cancelled')
        .values('employee_id', 'order_date')
        .annotate(paid=Sum('amount_paid_by_employee'))
    )
    emp_date_to_paid = defaultdict(lambda: Decimal('0'))
    for row in order_paid:
        order_date = row['order_date']
        if hasattr(order_date, 'date'):
            order_date = order_date.date()
        key = (row['employee_id'], order_date)
        emp_date_to_paid[key] = row['paid'] or Decimal('0')

    # Недели в диапазоне (пн = week_start)
    week_starts = set()
    d = start
    while d <= end:
        week_starts.add(d - timedelta(days=d.weekday()))
        d += timedelta(days=1)
    week_starts = sorted(week_starts)

    weeks_data = []
    remainder = Decimal('0')

    for week_index, week_start in enumerate(week_starts):
        week_end = week_start + timedelta(days=6)
        week_dates = [week_start + timedelta(days=i) for i in range(7)]
        week_dates_in_range = [d for d in week_dates if start <= d <= end]

        # Показываем всех сотрудников в таблице (даже если в неделе нет заказов).
        employees = list(Employee.objects.order_by('user__last_name', 'user__first_name'))

        table_rows = []
        totals_by_date = {d: {'fact': Decimal('0'), 'company': Decimal('0'), 'employee': Decimal('0')} for d in week_dates_in_range}
        total_company_week = Decimal('0')
        total_employee_week = Decimal('0')

        for num, emp in enumerate(employees, start=1):
            day_cells = []
            company_total = Decimal('0')
            employee_total = Decimal('0')
            for d in week_dates_in_range:
                fact_val = emp_date_to_fact.get((emp.id, d), Decimal('0'))
                paid_val = emp_date_to_paid.get((emp.id, d), Decimal('0'))
                if paid_val < 0:
                    paid_val = Decimal('0')
                if paid_val > fact_val:
                    paid_val = fact_val
                company_val = fact_val - paid_val
                total_company_week += company_val
                total_employee_week += paid_val
                company_total += company_val
                employee_total += paid_val
                day_cells.append({
                    'fact': fact_val,
                    'employee': paid_val,
                    'company': company_val,
                })
                totals_by_date[d]['fact'] += fact_val
                totals_by_date[d]['company'] += company_val
                totals_by_date[d]['employee'] += paid_val

            fact_total = sum(c['fact'] for c in day_cells)
            table_rows.append({
                'num': num,
                'name': emp.name,
                'personal_balance': emp.personal_balance,
                'company_total': company_total,
                'employee_total': employee_total,
                'day_cells': day_cells,
                'fact_total': fact_total,
            })

        total_day_cells = [totals_by_date[d] for d in week_dates_in_range]
        date_headers = [(d, WEEKDAY_NAMES_RU[d.weekday()]) for d in week_dates_in_range]

        wca = WeekCompanyAmount.objects.filter(week_start=week_start).first()
        company_given = (wca.amount if wca else Decimal('0'))
        remainder_prev = remainder
        available = company_given + remainder_prev
        remainder = available - total_company_week

        weeks_data.append({
            'week_start': week_start,
            'week_end': week_end,
            'week_number': week_index + 1,
            'week_dates': week_dates_in_range,
            'date_headers': date_headers,
            'table_rows': table_rows,
            'total_day_cells': total_day_cells,
            'company_given': company_given,
            'remainder_prev': remainder_prev,
            'total_company_week': total_company_week,
            'total_employee_week': total_employee_week,
            'total_fact_week': total_company_week + total_employee_week,
            'remainder_after': remainder,
        })

    return {
        'start_date': start.strftime('%Y-%m-%d'),
        'end_date': end.strftime('%Y-%m-%d'),
        'is_one_week': is_one_week,
        'weeks_data': weeks_data,
    }


def svodka_page(request):
    """Сводка заказов в формате таблицы Excel «ЭТУ ТАБЛИЦУ СМОТРИ»: по дням, колонки факт/компания/сотрудник."""
    context = _get_svodka_context(request)
    return render(request, 'orders/svodka.html', context)


def _build_svodka_xlsx(context):
    """Строит xlsx в формате листа «ЭТУ ТАБЛИЦУ СМОТРИ», возвращает bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Сводка'

    start = context['start_date']
    end = context['end_date']
    dates = context['dates']
    date_headers = context['date_headers']
    table_rows = context['table_rows']
    total_day_cells = context['total_day_cells']
    grand_total = context['grand_total_fact']

    # Строка 0: заголовок (писать в (1,1) до merge или после — после merge только (1,1) доступна)
    num_cols = 4 + 3 * len(dates) + 2
    ws.cell(1, 1, f'Траты на обеды с {start} по {end} года').alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    # Строка 1: даты (каждый день — 3 колонки); писать в первую ячейку диапазона до merge
    col = 5
    for d, day_name in date_headers:
        ws.cell(2, col, f"{d.strftime('%d.%m')} ({day_name})")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        col += 3

    # Строка 2: подзаголовки
    headers2 = [None, None, 'факт. оплата компании', 'факт. оплата сотрудником']
    for i, h in enumerate(headers2, 1):
        if h:
            ws.cell(3, i, h)
    col = 5
    for _ in dates:
        ws.cell(3, col, 'факт. стоимость ')
        ws.cell(3, col + 1, 'компания')
        ws.cell(3, col + 2, 'сотрудник')
        col += 3
    ws.cell(3, col, 'факт')
    ws.cell(3, col + 1, 'остаток (сотрудник)')

    # Данные
    for row_idx, r in enumerate(table_rows, start=4):
        ws.cell(row_idx, 1, r['num'])
        ws.cell(row_idx, 2, r['name'])
        ws.cell(row_idx, 3, float(r['fact_company_prepay']))
        ws.cell(row_idx, 4, float(r['fact_employee_prepay']))
        col = 5
        for cell in r['day_cells']:
            ws.cell(row_idx, col, float(cell['fact']))
            ws.cell(row_idx, col + 1, float(cell['company']))
            ws.cell(row_idx, col + 2, float(cell['employee']))
            col += 3
        ws.cell(row_idx, col, float(r['fact_total']))
        ws.cell(row_idx, col + 1, float(r['balance']))

    # Итого
    total_row = 4 + len(table_rows)
    ws.cell(total_row, 2, 'Итого:').font = Font(bold=True)
    ws.cell(total_row, 3, 0)
    ws.cell(total_row, 4, 0)
    col = 5
    for cell in total_day_cells:
        ws.cell(total_row, col, float(cell['fact']))
        ws.cell(total_row, col + 1, float(cell['company']))
        ws.cell(total_row, col + 2, float(cell['employee']))
        col += 3
    ws.cell(total_row, col, float(grand_total)).font = Font(bold=True)

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def svodka_mob_page(request):
    """Мобильная сводка: те же данные, адаптивная вёрстка + кнопка «Выгрузить в Excel»."""
    context = _get_svodka_context(request)
    if request.GET.get('export') == 'xlsx':
        xlsx_bytes = _build_svodka_xlsx(context)
        filename = f'svodka_{context["start_date"]}_{context["end_date"]}.xlsx'
        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return render(request, 'orders/svodka_mob.html', context)


def _build_svodka_mob2_xlsx(context):
    """Строит xlsx для svodka-mob2, возвращает bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Сводка моб2'

    start = context['start_date']
    end = context['end_date']
    weeks = context['weeks_data']

    row = 1
    ws.cell(row, 1, f'Финансовая сводка по неделям с {start} по {end}').font = Font(bold=True)
    row += 2

    for w in weeks:
        date_headers = w['date_headers']
        table_rows = w['table_rows']
        total_day_cells = w['total_day_cells']

        fixed_cols = 5  # №, сотрудник, личный, факт.компания, факт.сотрудник
        total_cols = fixed_cols + 3 * len(date_headers) + 1  # + Итого факт

        ws.cell(
            row,
            1,
            f"Неделя {w['week_number']} ({w['week_start'].strftime('%d.%m.%Y')} - {w['week_end'].strftime('%d.%m.%Y')})"
        ).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        row += 1

        # Верхняя шапка
        ws.cell(row, 1, '№')
        ws.cell(row, 2, 'Сотрудник')
        ws.cell(row, 3, 'Личный баланс')
        ws.cell(row, 4, 'факт. оплата компании')
        ws.cell(row, 5, 'факт. оплата сотрудником')
        col = 6
        for d, day_name in date_headers:
            ws.cell(row, col, f"{d.strftime('%d.%m')} ({day_name})")
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
            col += 3
        ws.cell(row, col, 'Итого факт')
        row += 1

        # Подшапка
        col = 6
        for _d, _day_name in date_headers:
            ws.cell(row, col, 'факт')
            ws.cell(row, col + 1, 'компания')
            ws.cell(row, col + 2, 'сотрудник')
            col += 3
        row += 1

        # Данные
        for r in table_rows:
            ws.cell(row, 1, r['num'])
            ws.cell(row, 2, r['name'])
            ws.cell(row, 3, float(r['personal_balance']))
            ws.cell(row, 4, float(r['company_total']))
            ws.cell(row, 5, float(r['employee_total']))
            col = 6
            for cell in r['day_cells']:
                ws.cell(row, col, float(cell['fact']))
                ws.cell(row, col + 1, float(cell['company']))
                ws.cell(row, col + 2, float(cell['employee']))
                col += 3
            ws.cell(row, col, float(r['fact_total']))
            row += 1

        # Итого
        ws.cell(row, 2, 'Итого:').font = Font(bold=True)
        ws.cell(row, 4, float(w['total_company_week'])).font = Font(bold=True)
        ws.cell(row, 5, float(w['total_employee_week'])).font = Font(bold=True)
        col = 6
        for cell in total_day_cells:
            ws.cell(row, col, float(cell['fact']))
            ws.cell(row, col + 1, float(cell['company']))
            ws.cell(row, col + 2, float(cell['employee']))
            col += 3
        ws.cell(row, col, float(w['total_fact_week'])).font = Font(bold=True)
        row += 1

        ws.cell(row, 1, 'Сколько дала компания на неделю:')
        ws.cell(row, 2, float(w['company_given']))
        ws.cell(row, 3, 'Остаток с прошлой недели:')
        ws.cell(row, 4, float(w['remainder_prev']))
        ws.cell(row, 5, 'Остаток после недели:')
        ws.cell(row, 6, float(w['remainder_after']))
        row += 2

    # Немного выравнивания
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=2):
        for c in r:
            c.alignment = Alignment(vertical='center')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def svodka_mob2_page(request):
    """Финансовая сводка по неделям: компания/сотрудник/факт по дням, «сколько дала компания», остаток с прошлой недели."""
    if request.method == 'POST':
        week_start_str = request.POST.get('week_start')
        amount_str = request.POST.get('amount', '0').strip()
        if week_start_str:
            try:
                week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
                amount = Decimal(amount_str) if amount_str else Decimal('0')
                WeekCompanyAmount.objects.update_or_create(
                    week_start=week_start,
                    defaults={'amount': amount}
                )
            except (ValueError, Exception):
                pass
        redirect_url = request.path + '?' + request.GET.urlencode() if request.GET else request.path
        return redirect(redirect_url)

    context = _get_svodka_mob2_context(request)
    if request.GET.get('export') == 'xlsx':
        xlsx_bytes = _build_svodka_mob2_xlsx(context)
        filename = f'svodka_mob2_{context["start_date"]}_{context["end_date"]}.xlsx'
        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return render(request, 'orders/svodka_mob2.html', context)


def svodka_mob2_help_page(request):
    """Справка: как считается каждое поле на странице svodka-mob2."""
    return render(request, 'orders/svodka_mob2_help.html')


@login_required
@user_passes_test(is_superuser)
def backup_database(request):
    """Выгрузка бэкапа базы данных в виде фикстур"""
    from django.core.management import call_command
    from io import StringIO
    import time
    
    try:
        output = StringIO()
        # Выгружаем все данные приложения orders
        call_command('dumpdata', 'orders', indent=2, stdout=output)
        
        # Сохраняем в статику
        filename = f'db_backup_{int(time.time())}.json'
        backup_path = os.path.join(settings.STATIC_ROOT, 'backups', filename)
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)
        
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(output.getvalue())
            
        # Возвращаем файл для скачивания
        response = HttpResponse(output.getvalue(), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return HttpResponse(f'Error creating backup: {str(e)}', status=500)
